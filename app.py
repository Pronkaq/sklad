from __future__ import annotations

import io
import secrets
import sqlite3
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import qrcode
import qrcode.image.svg
from functools import wraps
from flask import Flask, Response, redirect, render_template, request, send_file, url_for, session
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


APP_DIR = Path(__file__).resolve().parent
DB_PATH = APP_DIR / "pallets.db"

app = Flask(__name__)
app.secret_key = "change-this-secret-key-in-production"


STATUSES = {
    "CREATED": "Создан у источника",
    "IN_SHOP5_WAREHOUSE": "На цеховом складе 5",
    "IN_PROCESSING": "В обработке",
    "PROCESSED": "Обработан",
    "IN_INSPECTION": "На разбраковке",
    "TO": "ТО",
    "WAREHOUSE": "Цеховой склад 5",
    "FG_WAREHOUSE": "Склад готовой продукции",
    "SHOP8_WAREHOUSE": "Цех 8",
    "LAB7": "7 лаборатория",
    "REJECT": "Брак",
    "CONDITIONALLY_OK": "Условно годен",
    "SOLD": "Продано",
    "SOLD": "Продано",
    "CLOSED": "Закрыт / разделён",
}

LOCATIONS = {
    "SHOP10": "Цех 10",
    "SHOP3": "Цех 3",
    "NOVGOROD": "Новгород",
    "SHOP5_WAREHOUSE": "Цеховой склад 5",
    "FG_WAREHOUSE": "Склад готовой продукции",
    "SHOP8_WAREHOUSE": "Цех 8",
    "KTU1": "КТУ-1",
    "KTU2": "КТУ-2",
    "LKTM": "ЛКТМ",
    "INSPECTION": "Разбраковка",
    "TO": "ТО",
    "WAREHOUSE": "Цеховой склад 5",
    "LAB7": "7 лаборатория",
    "REJECT": "Брак",
    "CONDITIONALLY_OK": "Условно годен",
}

DESTINATIONS = [
    ("TO", "ТО"),
    ("WAREHOUSE", "Цеховой склад 5"),
    ("FG_WAREHOUSE", "Склад готовой продукции"),
    ("SHOP8_WAREHOUSE", "Цех 8"),
    ("LAB7", "7 лаборатория"),
    ("REJECT", "Брак"),
    ("CONDITIONALLY_OK", "Условно годен"),
]


ASSORTMENTS = [
    "СТ-11-13",
    "11С6-8/3-94(130)",
    "11С6-8/3-94(94)",
    "ТС-8/3-К",
    "ТР-11-1,1-13",
]

ASSORTMENTS_PROCESSED = [
    "КТ-11-13",
    "11С6-8/3-94(130)",
    "11С6-8/3-94(94)",
    "ТС-8/3-К",
    "ТР-11-1,1-13",
]

ITEM_CATEGORIES = [
    ("fabric", "Ткань"),
    ("thread", "Нить"),
    ("cord", "Шнур"),
    ("tape", "Лента"),
    ("other", "Другое"),
]

DEFAULT_USERS = [
    # login, password, display_name, shop, role
    # source_user/source_admin: только создание этикетки передачи в цех 5
    ("ceh10", "ceh10", "Цех 10", "SHOP10", "source_user"),
    ("adminceh10", "adminceh10", "Админ цех 10", "SHOP10", "source_admin"),
    ("ceh3", "ceh3", "Цех 3", "SHOP3", "source_user"),
    ("adminceh3", "adminceh3", "Админ цех 3", "SHOP3", "source_admin"),
    ("novgorod", "novgorod", "Новгород", "NOVGOROD", "source_user"),
    ("adminnovgorod", "adminnovgorod", "Админ Новгород", "NOVGOROD", "source_admin"),

    # shop5_user/shop5_admin: приёмка, обработка, разбраковка, остатки, Excel
    ("ceh5", "ceh5", "Цех 5", "SHOP5", "shop5_user"),
    ("adminceh5", "adminceh5", "Админ цех 5", "SHOP5", "shop5_admin"),
]


def now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def parse_float(value: Any, default: float = 0.0) -> float:
    if value is None:
        return default

    normalized = str(value).strip().replace(",", ".")
    if normalized == "":
        return default

    try:
        return float(normalized)
    except (TypeError, ValueError):
        return default


def get_db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS pallets (
            id TEXT PRIMARY KEY,
            parent_id TEXT,
            source_shop TEXT NOT NULL,
            assortment TEXT NOT NULL,
            party_number TEXT,
            item_category TEXT DEFAULT 'fabric',
            created_reason TEXT DEFAULT 'production_transfer',
            rolls_count INTEGER NOT NULL,
            meters_total REAL NOT NULL,
            production_date TEXT,
            processing_date TEXT,
            process_type TEXT,
            machine TEXT,
            responsible TEXT,
            status TEXT NOT NULL,
            location TEXT NOT NULL,
            created_by TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            comment TEXT
        )
    """)

    try:
        cur.execute("ALTER TABLE pallets ADD COLUMN party_number TEXT")
    except sqlite3.OperationalError:
        pass

    try:
        cur.execute("ALTER TABLE pallets ADD COLUMN item_category TEXT DEFAULT 'fabric'")
    except sqlite3.OperationalError:
        pass

    try:
        cur.execute("ALTER TABLE pallets ADD COLUMN created_reason TEXT DEFAULT 'production_transfer'")
    except sqlite3.OperationalError:
        pass

    cur.execute("""
        CREATE TABLE IF NOT EXISTS assortments (
            name TEXT PRIMARY KEY,
            category TEXT NOT NULL DEFAULT 'fabric',
            active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL
        )
    """)

    for name in ASSORTMENTS:
        cur.execute("""
            INSERT OR IGNORE INTO assortments(name, category, active, created_at)
            VALUES (?, 'fabric', 1, ?)
        """, (name, now_str()))

    for name in ASSORTMENTS_PROCESSED:
        cur.execute("""
            INSERT OR IGNORE INTO assortments(name, category, active, created_at)
            VALUES (?, 'fabric', 1, ?)
        """, (name, now_str()))

    cur.execute("""
        CREATE TABLE IF NOT EXISTS movements (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pallet_id TEXT NOT NULL,
            action TEXT NOT NULL,
            from_status TEXT,
            to_status TEXT,
            from_location TEXT,
            to_location TEXT,
            user TEXT,
            timestamp TEXT NOT NULL,
            comment TEXT
        )
    """)
    cur.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS idx_unique_shop5_accept
        ON movements (pallet_id)
        WHERE action = 'Приёмка в цех 5'
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            username TEXT PRIMARY KEY,
            password TEXT NOT NULL,
            display_name TEXT NOT NULL,
            shop TEXT NOT NULL,
            role TEXT NOT NULL,
            active INTEGER NOT NULL DEFAULT 1
        )
    """)

    for username, password, display_name, shop, role in DEFAULT_USERS:
        cur.execute("""
            INSERT OR IGNORE INTO users(username, password, display_name, shop, role, active)
            VALUES (?, ?, ?, ?, ?, 1)
        """, (username, password, display_name, shop, role))

    cur.execute("""
        CREATE TABLE IF NOT EXISTS counters (
            year INTEGER PRIMARY KEY,
            value INTEGER NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS counters_scoped (
            counter_key TEXT NOT NULL,
            year INTEGER NOT NULL,
            value INTEGER NOT NULL,
            PRIMARY KEY (counter_key, year)
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS roll_labels (
            id TEXT PRIMARY KEY,
            source_shop TEXT NOT NULL,
            assortment TEXT NOT NULL,
            roll_number TEXT NOT NULL,
            party_number TEXT,
            base_number TEXT,
            lubricant TEXT,
            width TEXT,
            meters REAL NOT NULL,
            meters_shift_1 REAL,
            meters_shift_2 REAL,
            weaver_name TEXT,
            weaver_name_2 TEXT,
            loom_number INTEGER,
            assistant_name TEXT,
            production_date TEXT,
            created_by TEXT,
            created_at TEXT NOT NULL,
            pallet_id TEXT,
            comment TEXT
        )
    """)
    try:
        cur.execute("ALTER TABLE roll_labels ADD COLUMN weaver_name_2 TEXT")
    except sqlite3.OperationalError:
        pass
    try:
        cur.execute("ALTER TABLE roll_labels ADD COLUMN meters_shift_1 REAL")
    except sqlite3.OperationalError:
        pass
    try:
        cur.execute("ALTER TABLE roll_labels ADD COLUMN meters_shift_2 REAL")
    except sqlite3.OperationalError:
        pass
    try:
        cur.execute("ALTER TABLE roll_labels ADD COLUMN loom_number INTEGER")
    except sqlite3.OperationalError:
        pass

    cur.execute("""
        CREATE TABLE IF NOT EXISTS shop10_staff (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            role TEXT NOT NULL CHECK(role IN ('weaver', 'assistant')),
            full_name TEXT NOT NULL,
            active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL,
            UNIQUE(role, full_name)
        )
    """)

    # Data-consistency repair: unlink rolls from deleted/non-existent pallets.
    cur.execute("""
        UPDATE roll_labels
        SET pallet_id = NULL
        WHERE pallet_id IS NOT NULL
          AND pallet_id NOT IN (SELECT id FROM pallets)
    """)

    conn.commit()
    conn.close()


def _get_max_existing_counter(conn: sqlite3.Connection, prefix: str, year: int) -> int:
    pattern = f"{prefix}-{year}-%"
    if prefix == "RL":
        row = conn.execute(
            """
            SELECT MAX(CAST(substr(id, -6) AS INTEGER)) AS max_value
            FROM roll_labels
            WHERE id LIKE ?
            """,
            (pattern,),
        ).fetchone()
    else:
        row = conn.execute(
            """
            SELECT MAX(CAST(substr(id, -6) AS INTEGER)) AS max_value
            FROM pallets
            WHERE id LIKE ?
            """,
            (pattern,),
        ).fetchone()
    return int(row["max_value"] or 0)


def _next_scoped_id(conn: sqlite3.Connection, prefix: str = "P") -> str:
    year = datetime.now().year
    cur = conn.cursor()
    counter_key = prefix.strip().upper() or "P"
    cur.execute(
        "SELECT value FROM counters_scoped WHERE counter_key = ? AND year = ?",
        (counter_key, year),
    )
    row = cur.fetchone()
    if row is None:
        value = _get_max_existing_counter(conn, counter_key, year) + 1
        cur.execute(
            "INSERT INTO counters_scoped(counter_key, year, value) VALUES(?, ?, ?)",
            (counter_key, year, value),
        )
    else:
        value = row["value"] + 1
        cur.execute(
            "UPDATE counters_scoped SET value = ? WHERE counter_key = ? AND year = ?",
            (value, counter_key, year),
        )
    return f"{counter_key}-{year}-{value:06d}"


def next_pallet_id(prefix: str = "P") -> str:
    conn = get_db()
    value = _next_scoped_id(conn, prefix)
    conn.commit()
    conn.close()
    return value


def add_movement(
    conn: sqlite3.Connection,
    pallet_id: str,
    action: str,
    from_status: str | None,
    to_status: str | None,
    from_location: str | None,
    to_location: str | None,
    user: str | None,
    comment: str | None = None,
) -> None:
    conn.execute("""
        INSERT INTO movements(
            pallet_id, action, from_status, to_status,
            from_location, to_location, user, timestamp, comment
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        pallet_id, action, from_status, to_status,
        from_location, to_location, user, now_str(), comment
    ))


def get_assortments(category: str | None = None) -> list[sqlite3.Row]:
    conn = get_db()
    if category:
        rows = conn.execute("""
            SELECT * FROM assortments
            WHERE active = 1 AND category = ?
            ORDER BY name
        """, (category,)).fetchall()
    else:
        rows = conn.execute("""
            SELECT * FROM assortments
            WHERE active = 1
            ORDER BY category, name
        """).fetchall()
    conn.close()
    return rows




def next_pallet_id_tx(conn: sqlite3.Connection, prefix: str = "P") -> str:
    return _next_scoped_id(conn, prefix)

def get_pallet_or_404(pallet_id: str) -> sqlite3.Row:
    conn = get_db()
    pallet = conn.execute("SELECT * FROM pallets WHERE id = ?", (pallet_id,)).fetchone()
    conn.close()
    if pallet is None:
        from flask import abort
        abort(404)
    return pallet


def can_delete_pallet(user: sqlite3.Row | None, pallet: sqlite3.Row) -> bool:
    if not user:
        return False
    if is_shop5_admin(user):
        return True
    if is_source_user(user):
        return pallet["source_shop"] == user["shop"] and pallet["status"] == "CREATED"
    return False


def can_edit_shop10_pallet_label(user: sqlite3.Row | None, pallet: sqlite3.Row | None) -> bool:
    if not user or not pallet:
        return False
    if not is_shop10_user(user):
        return False
    return pallet["source_shop"] == user["shop"] and pallet["status"] not in ("CLOSED", "SOLD")


def can_edit_manual_pallet_label(user: sqlite3.Row | None, pallet: sqlite3.Row | None) -> bool:
    if not user or not pallet:
        return False
    if not is_shop5_user(user):
        return False
    return (
        (pallet["created_reason"] or "production_transfer") == "manual_inventory"
        and pallet["status"] not in ("CLOSED", "SOLD")
    )


def manual_location_status(location: str) -> str:
    if location == "SHOP5_WAREHOUSE":
        return "IN_SHOP5_WAREHOUSE"
    if location == "FG_WAREHOUSE":
        return "FG_WAREHOUSE"
    if location == "INSPECTION":
        return "IN_INSPECTION"
    if location in ("KTU1", "KTU2", "LKTM"):
        return "IN_PROCESSING"
    return location



def generate_csrf_token() -> str:
    token = session.get("csrf_token")
    if not token:
        token = secrets.token_urlsafe(32)
        session["csrf_token"] = token
    return token


def verify_csrf() -> None:
    if request.method != "POST":
        return

    session_token = session.get("csrf_token")
    form_token = request.form.get("csrf_token", "")
    if not session_token or not form_token or form_token != session_token:
        from flask import abort
        abort(400, description="CSRF token missing or invalid")


@app.before_request
def csrf_protect() -> None:
    verify_csrf()

def current_user() -> sqlite3.Row | None:
    username = session.get("username")
    if not username:
        return None

    conn = get_db()
    user = conn.execute(
        "SELECT * FROM users WHERE username = ? AND active = 1",
        (username,)
    ).fetchone()
    conn.close()
    return user


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("username"):
            return redirect(url_for("login", next=request.path))
        return view(*args, **kwargs)
    return wrapped


def is_source_user(user: sqlite3.Row | None) -> bool:
    return bool(user and user["role"] in ("source_user", "source_admin"))


def is_source_admin(user: sqlite3.Row | None) -> bool:
    return bool(user and user["role"] == "source_admin")


def is_shop5_user(user: sqlite3.Row | None) -> bool:
    return bool(user and user["role"] in ("shop5_user", "shop5_admin"))


def is_shop5_admin(user: sqlite3.Row | None) -> bool:
    return bool(user and user["role"] == "shop5_admin")


def is_shop10_user(user: sqlite3.Row | None) -> bool:
    return bool(user and user["shop"] == "SHOP10")


def can_view_roll_label(user: sqlite3.Row | None, roll: sqlite3.Row | None) -> bool:
    if not user or not roll:
        return False
    return bool(is_shop5_user(user) or roll["source_shop"] == user["shop"])


def can_manage_roll_label(user: sqlite3.Row | None, roll: sqlite3.Row | None) -> bool:
    if not user or not roll:
        return False
    if not is_shop10_user(user):
        return False
    if roll["source_shop"] != user["shop"]:
        return False
    if roll["pallet_id"]:
        return False
    try:
        created_at = datetime.strptime(roll["created_at"], "%Y-%m-%d %H:%M:%S")
    except Exception:
        return False
    return (datetime.now() - created_at) <= timedelta(days=1)


def shop5_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        user = current_user()
        if not user:
            return redirect(url_for("login", next=request.path))
        if not is_shop5_user(user):
            return render_template(
                "access_denied.html",
                message="Эта операция доступна только цеху 5."
            ), 403
        return view(*args, **kwargs)
    return wrapped


@app.route("/login", methods=["GET", "POST"])
def login():
    error = None

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        conn = get_db()
        user = conn.execute("""
            SELECT * FROM users
            WHERE username = ? AND password = ? AND active = 1
        """, (username, password)).fetchone()
        conn.close()

        if user:
            session["username"] = user["username"]
            return redirect(request.args.get("next") or url_for("index"))

        error = "Неверный логин или пароль"

    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.context_processor
def inject_helpers() -> dict[str, Any]:
    return {
        "STATUSES": STATUSES,
        "LOCATIONS": LOCATIONS,
        "ASSORTMENTS": ASSORTMENTS,
        "ASSORTMENTS_PROCESSED": ASSORTMENTS_PROCESSED,
        "ITEM_CATEGORIES": ITEM_CATEGORIES,
        "db_assortments": get_assortments(),
        "current_user": current_user(),
        "is_source_user": is_source_user,
        "is_source_admin": is_source_admin,
        "is_shop5_user": is_shop5_user,
        "is_shop5_admin": is_shop5_admin,
        "is_shop10_user": is_shop10_user,
        "can_manage_roll_label": can_manage_roll_label,
        "can_delete_pallet": can_delete_pallet,
        "can_edit_shop10_pallet_label": can_edit_shop10_pallet_label,
        "can_edit_manual_pallet_label": can_edit_manual_pallet_label,
        "csrf_token": generate_csrf_token,
    }


@app.route("/")
@login_required
def index():
    user = current_user()
    conn = get_db()

    if is_source_user(user):
        pallets = conn.execute("""
            SELECT * FROM pallets
            WHERE source_shop = ?
              AND parent_id IS NULL
              AND COALESCE(created_reason, 'production_transfer') = 'production_transfer'
            ORDER BY created_at DESC
            LIMIT 100
        """, (user["shop"],)).fetchall()

        stats = conn.execute("""
            SELECT source_shop, assortment,
                   COUNT(*) AS pallets_count,
                   SUM(rolls_count) AS rolls_sum,
                   SUM(meters_total) AS meters_sum
            FROM pallets
            WHERE source_shop = ?
              AND parent_id IS NULL
              AND COALESCE(created_reason, 'production_transfer') = 'production_transfer'
            GROUP BY source_shop, assortment
            ORDER BY assortment
        """, (user["shop"],)).fetchall()

        monthly_stats = conn.execute("""
            SELECT substr(created_at, 1, 7) AS month,
                   assortment,
                   COUNT(*) AS pallets_count,
                   SUM(rolls_count) AS rolls_sum,
                   SUM(meters_total) AS meters_sum
            FROM pallets
            WHERE source_shop = ?
              AND parent_id IS NULL
              AND COALESCE(created_reason, 'production_transfer') = 'production_transfer'
            GROUP BY substr(created_at, 1, 7), assortment
            ORDER BY month DESC, assortment
            LIMIT 100
        """, (user["shop"],)).fetchall()
    else:
        pallets = conn.execute("""
            SELECT * FROM pallets
            ORDER BY updated_at DESC
            LIMIT 100
        """).fetchall()

        stats = conn.execute("""
            SELECT status, location, assortment,
                   COUNT(*) AS pallets_count,
                   SUM(rolls_count) AS rolls_sum,
                   SUM(meters_total) AS meters_sum
            FROM pallets
            WHERE status NOT IN ('CLOSED', 'SOLD')
            GROUP BY status, location, assortment
            ORDER BY status, location, assortment
        """).fetchall()

        monthly_stats = []

    conn.close()
    return render_template("index.html", pallets=pallets, stats=stats, monthly_stats=monthly_stats)


@app.route("/pallet/new", methods=["GET", "POST"])
@login_required
def create_pallet():
    user = current_user()
    if is_shop5_user(user):
        return render_template(
            "access_denied.html",
            message="Цех 5 не создаёт этикетки передачи от цехов 10/3/Новгород. Цех 5 принимает уже созданные поддоны через карточку/сканирование."
        ), 403

    if request.method == "POST":
        if user and user["role"] in ("source_user",):
            source_shop = user["shop"]
        else:
            source_shop = request.form["source_shop"].strip()
        assortment = request.form["assortment"].strip()
        party_number = request.form.get("party_number", "").strip()
        item_category = request.form.get("item_category", "fabric").strip()
        rolls_count = int(request.form["rolls_count"])
        meters_total = float(request.form["meters_total"])
        production_date = request.form.get("production_date") or ""
        responsible = request.form.get("responsible", "").strip() or (current_user()["display_name"] if current_user() else "")
        comment = request.form.get("comment", "").strip()

        pallet_id = next_pallet_id("P")
        conn = get_db()
        conn.execute("""
            INSERT INTO pallets(
                id, parent_id, source_shop, assortment, party_number, item_category, created_reason, rolls_count, meters_total,
                production_date, processing_date, process_type, machine,
                responsible, status, location, created_by,
                created_at, updated_at, comment
            )
            VALUES (?, NULL, ?, ?, ?, ?, 'production_transfer', ?, ?, ?, NULL, NULL, NULL, ?, 'CREATED', ?, ?, ?, ?, ?)
        """, (
            pallet_id, source_shop, assortment, party_number, item_category, rolls_count, meters_total,
            production_date, responsible, source_shop, responsible,
            now_str(), now_str(), comment
        ))

        add_movement(
            conn, pallet_id, "Создание поддона",
            None, "CREATED", None, source_shop,
            responsible, comment
        )

        conn.commit()
        conn.close()
        return redirect(url_for("pallet_detail", pallet_id=pallet_id))

    return render_template("create_pallet.html", user=current_user())


@app.route("/exp/shop10/pallet-labels/<pallet_id>/edit", methods=["GET", "POST"])
@login_required
def edit_shop10_pallet_label(pallet_id: str):
    user = current_user()
    conn = get_db()
    pallet = conn.execute("SELECT * FROM pallets WHERE id = ?", (pallet_id,)).fetchone()
    if not can_edit_shop10_pallet_label(user, pallet):
        conn.close()
        return render_template("access_denied.html", message="Редактирование доступно только для активных этикеток цеха 10 (кроме CLOSED/SOLD)."), 403

    if request.method == "POST":
        assortment = request.form["assortment"].strip()
        party_number = request.form.get("party_number", "").strip()
        item_category = request.form.get("item_category", "fabric").strip()
        rolls_count = int(request.form["rolls_count"])
        meters_total = float(request.form["meters_total"])
        production_date = request.form.get("production_date") or ""
        responsible = request.form.get("responsible", "").strip() or (user["display_name"] if user else "")
        comment = request.form.get("comment", "").strip()

        conn.execute(
            """
            UPDATE pallets
            SET assortment = ?,
                party_number = ?,
                item_category = ?,
                rolls_count = ?,
                meters_total = ?,
                production_date = ?,
                responsible = ?,
                comment = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (assortment, party_number, item_category, rolls_count, meters_total, production_date, responsible, comment, now_str(), pallet_id),
        )
        conn.commit()
        conn.close()
        return redirect(url_for("pallet_detail", pallet_id=pallet_id))

    conn.close()
    return render_template("create_pallet.html", user=user, edit_pallet=pallet)


@app.route("/pallet/manual", methods=["GET", "POST"])
@shop5_required
def manual_pallet():
    if request.method == "POST":
        assortment = request.form["assortment"].strip()
        item_category = request.form.get("item_category", "fabric").strip()
        party_number = request.form.get("party_number", "").strip()
        rolls_count = int(request.form.get("rolls_count") or 0)
        meters_total = parse_float(request.form.get("meters_total") or 0)
        location = request.form.get("location", "SHOP5_WAREHOUSE").strip()
        responsible = request.form.get("responsible", "").strip() or (current_user()["display_name"] if current_user() else "")
        comment = request.form.get("comment", "").strip()

        status = manual_location_status(location)

        pallet_id = next_pallet_id("INV")
        conn = get_db()

        conn.execute("""
            INSERT INTO pallets(
                id, parent_id, source_shop, assortment, party_number, item_category, created_reason,
                rolls_count, meters_total, production_date, processing_date, process_type, machine,
                responsible, status, location, created_by,
                created_at, updated_at, comment
            )
            VALUES (?, NULL, ?, ?, ?, ?, 'manual_inventory',
                    ?, ?, NULL, NULL, NULL, NULL,
                    ?, ?, ?, ?, ?, ?, ?)
        """, (
            pallet_id, location, assortment, party_number, item_category,
            rolls_count, meters_total,
            responsible, status, location, responsible,
            now_str(), now_str(), comment
        ))

        add_movement(
            conn, pallet_id, "Создана этикетка текущего остатка",
            None, status, None, location, responsible, comment
        )

        conn.commit()
        conn.close()
        return redirect(url_for("pallet_detail", pallet_id=pallet_id))

    return render_template("manual_pallet.html", assortments=get_assortments())


@app.route("/pallet/manual/<pallet_id>/edit", methods=["GET", "POST"])
@shop5_required
def edit_manual_pallet(pallet_id: str):
    user = current_user()
    conn = get_db()
    pallet = conn.execute("SELECT * FROM pallets WHERE id = ?", (pallet_id,)).fetchone()
    if not can_edit_manual_pallet_label(user, pallet):
        conn.close()
        return render_template(
            "access_denied.html",
            message="Редактирование доступно только для активных этикеток текущего остатка цеха 5.",
        ), 403

    if request.method == "POST":
        assortment = request.form["assortment"].strip()
        item_category = request.form.get("item_category", "fabric").strip()
        party_number = request.form.get("party_number", "").strip()
        rolls_count = int(request.form.get("rolls_count") or 0)
        meters_total = parse_float(request.form.get("meters_total") or 0)
        location = request.form.get("location", pallet["location"]).strip()
        status = manual_location_status(location)
        responsible = request.form.get("responsible", "").strip() or (user["display_name"] if user else "")
        comment = request.form.get("comment", "").strip()

        conn.execute(
            """
            UPDATE pallets
            SET source_shop = ?,
                assortment = ?,
                party_number = ?,
                item_category = ?,
                rolls_count = ?,
                meters_total = ?,
                responsible = ?,
                status = ?,
                location = ?,
                comment = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (
                location, assortment, party_number, item_category,
                rolls_count, meters_total, responsible, status, location,
                comment, now_str(), pallet_id,
            ),
        )

        add_movement(
            conn, pallet_id, "Редактирование этикетки текущего остатка",
            pallet["status"], status, pallet["location"], location,
            responsible, comment,
        )

        conn.commit()
        conn.close()
        return redirect(url_for("pallet_detail", pallet_id=pallet_id))

    conn.close()
    return render_template("manual_pallet.html", assortments=get_assortments(), edit_pallet=pallet)


@app.route("/pallet/<pallet_id>")
@login_required
def pallet_detail(pallet_id: str):
    conn = get_db()
    pallet = conn.execute("SELECT * FROM pallets WHERE id = ?", (pallet_id,)).fetchone()
    pallet_rolls = conn.execute("""
        SELECT id, roll_number, assortment, party_number, meters, production_date
        FROM roll_labels
        WHERE pallet_id = ?
        ORDER BY id ASC
    """, (pallet_id,)).fetchall()
    movements = conn.execute("""
        SELECT * FROM movements
        WHERE pallet_id = ?
        ORDER BY timestamp DESC, id DESC
    """, (pallet_id,)).fetchall()

    children = conn.execute("""
        SELECT * FROM pallets
        WHERE parent_id = ?
        ORDER BY created_at DESC
    """, (pallet_id,)).fetchall()
    shop5_accept_exists = conn.execute("""
        SELECT 1
        FROM movements
        WHERE pallet_id = ? AND action = 'Приёмка в цех 5'
        LIMIT 1
    """, (pallet_id,)).fetchone() is not None

    conn.close()

    if pallet is None:
        from flask import abort
        abort(404)

    return render_template(
        "pallet_detail.html",
        pallet=pallet,
        pallet_rolls=pallet_rolls,
        movements=movements,
        children=children,
        destinations=DESTINATIONS,
        shop5_accept_exists=shop5_accept_exists,
    )


@app.route("/pallet/<pallet_id>/accept", methods=["POST"])
@shop5_required
def accept_to_shop5(pallet_id: str):
    user = request.form.get("user", "").strip()
    pallet = get_pallet_or_404(pallet_id)
    if pallet["status"] != "CREATED":
        return redirect(url_for("pallet_detail", pallet_id=pallet_id))

    conn = get_db()
    conn.execute("BEGIN IMMEDIATE")
    existing_accept = conn.execute("""
        SELECT 1
        FROM movements
        WHERE pallet_id = ? AND action = 'Приёмка в цех 5'
        LIMIT 1
    """, (pallet_id,)).fetchone()
    if not existing_accept:
        conn.execute("""
            UPDATE pallets
            SET status = 'IN_SHOP5_WAREHOUSE',
                location = 'SHOP5_WAREHOUSE',
                updated_at = ?
            WHERE id = ?
        """, (now_str(), pallet_id))

        add_movement(
            conn, pallet_id, "Приёмка в цех 5",
            pallet["status"], "IN_SHOP5_WAREHOUSE",
            pallet["location"], "SHOP5_WAREHOUSE",
            user
        )

    conn.commit()
    conn.close()
    return redirect(url_for("pallet_detail", pallet_id=pallet_id))


@app.route("/pallet/<pallet_id>/delete", methods=["POST"])
@login_required
def delete_pallet(pallet_id: str):
    user = current_user()
    pallet = get_pallet_or_404(pallet_id)
    if not can_delete_pallet(user, pallet):
        return render_template(
            "access_denied.html",
            message="Удаление этой этикетки вам недоступно."
        ), 403

    conn = get_db()
    conn.execute("UPDATE roll_labels SET pallet_id = NULL WHERE pallet_id = ?", (pallet_id,))
    conn.execute("DELETE FROM movements WHERE pallet_id = ?", (pallet_id,))
    conn.execute("DELETE FROM pallets WHERE id = ?", (pallet_id,))
    conn.commit()
    conn.close()
    return redirect(url_for("index"))


@app.route("/pallet/<pallet_id>/send-processing", methods=["POST"])
@shop5_required
def send_processing(pallet_id: str):
    user = request.form.get("user", "").strip()
    machine = request.form.get("machine", "").strip()
    process_type = request.form.get("process_type", "").strip()
    pallet = get_pallet_or_404(pallet_id)

    conn = get_db()
    conn.execute("""
        UPDATE pallets
        SET status = 'IN_PROCESSING',
            location = ?,
            machine = ?,
            process_type = ?,
            updated_at = ?
        WHERE id = ?
    """, (machine, machine, process_type, now_str(), pallet_id))

    add_movement(
        conn, pallet_id, "Передача на обработку",
        pallet["status"], "IN_PROCESSING",
        pallet["location"], machine,
        user, process_type
    )

    conn.commit()
    conn.close()
    return redirect(url_for("pallet_detail", pallet_id=pallet_id))


@app.route("/pallet/<pallet_id>/finish-processing", methods=["POST"])
@shop5_required
def finish_processing(pallet_id: str):
    pallet = get_pallet_or_404(pallet_id)

    assortment = request.form.get("assortment", pallet["assortment"]).strip()
    rolls_count = int(request.form.get("rolls_count", pallet["rolls_count"]))
    meters_total = float(request.form.get("meters_total", pallet["meters_total"]))
    processing_date = request.form.get("processing_date") or datetime.now().strftime("%Y-%m-%d")
    party_number = request.form.get("party_number", pallet["party_number"] or "").strip()
    process_type = request.form.get("process_type", pallet["process_type"] or "").strip()
    machine = request.form.get("machine", pallet["machine"] or "").strip()
    responsible = request.form.get("responsible", "").strip()
    comment = request.form.get("comment", "").strip()

    new_id = next_pallet_id("OBR")

    conn = get_db()

    conn.execute("""
        UPDATE pallets
        SET status = 'CLOSED',
            updated_at = ?
        WHERE id = ?
    """, (now_str(), pallet_id))

    add_movement(
        conn, pallet_id, "Закрыт после обработки, создан новый обработанный поддон",
        pallet["status"], "CLOSED",
        pallet["location"], pallet["location"],
        responsible, f"Новый поддон: {new_id}"
    )

    conn.execute("""
        INSERT INTO pallets(
            id, parent_id, source_shop, assortment, party_number, item_category, created_reason, rolls_count, meters_total,
            production_date, processing_date, process_type, machine,
            responsible, status, location, created_by,
            created_at, updated_at, comment
        )
        VALUES (?, ?, ?, ?, ?, ?, 'processing', ?, ?, ?, ?, ?, ?, ?, 'PROCESSED', ?, ?, ?, ?, ?)
    """, (
        new_id, pallet_id, pallet["source_shop"], assortment, party_number, pallet["item_category"] or "fabric", rolls_count, meters_total,
        pallet["production_date"], processing_date, process_type, machine,
        responsible, machine or pallet["location"], responsible,
        now_str(), now_str(), comment
    ))

    add_movement(
        conn, new_id, "Создан обработанный поддон",
        None, "PROCESSED",
        None, machine or pallet["location"],
        responsible, f"Исходный поддон: {pallet_id}"
    )

    conn.commit()
    conn.close()
    return redirect(url_for("pallet_detail", pallet_id=new_id))


@app.route("/pallet/<pallet_id>/start-inspection", methods=["POST"])
@shop5_required
def start_inspection(pallet_id: str):
    user = request.form.get("user", "").strip()
    pallet = get_pallet_or_404(pallet_id)

    conn = get_db()
    conn.execute("""
        UPDATE pallets
        SET status = 'IN_INSPECTION',
            location = 'INSPECTION',
            updated_at = ?
        WHERE id = ?
    """, (now_str(), pallet_id))

    add_movement(
        conn, pallet_id, "Передача на разбраковку",
        pallet["status"], "IN_INSPECTION",
        pallet["location"], "INSPECTION",
        user
    )

    conn.commit()
    conn.close()
    return redirect(url_for("pallet_detail", pallet_id=pallet_id))


@app.route("/pallet/<pallet_id>/split-inspection", methods=["POST"])
@shop5_required
def split_inspection(pallet_id: str):
    pallet = get_pallet_or_404(pallet_id)
    user = request.form.get("user", "").strip()
    comment = request.form.get("comment", "").strip()

    conn = get_db()

    created_ids = []

    for dest_code, dest_label in DESTINATIONS:
        rolls_raw = request.form.get(f"rolls_{dest_code}", "").strip()
        meters_raw = request.form.get(f"meters_{dest_code}", "").strip()

        if not rolls_raw and not meters_raw:
            continue

        rolls = int(rolls_raw or 0)
        meters = float(meters_raw or 0)

        if rolls <= 0 and meters <= 0:
            continue

        new_id = next_pallet_id(dest_code)

        conn.execute("""
            INSERT INTO pallets(
                id, parent_id, source_shop, assortment, party_number, item_category, created_reason, rolls_count, meters_total,
                production_date, processing_date, process_type, machine,
                responsible, status, location, created_by,
                created_at, updated_at, comment
            )
            VALUES (?, ?, ?, ?, ?, ?, 'inspection_split', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            new_id, pallet_id, pallet["source_shop"], pallet["assortment"], pallet["party_number"], pallet["item_category"] or "fabric", rolls, meters,
            pallet["production_date"], pallet["processing_date"],
            pallet["process_type"], pallet["machine"],
            user, dest_code, dest_code, user,
            now_str(), now_str(), comment
        ))

        add_movement(
            conn, new_id, f"Создано после разбраковки: {dest_label}",
            None, dest_code,
            None, dest_code,
            user, f"Исходный поддон: {pallet_id}"
        )

        created_ids.append(new_id)

    conn.execute("""
        UPDATE pallets
        SET status = 'CLOSED',
            updated_at = ?
        WHERE id = ?
    """, (now_str(), pallet_id))

    add_movement(
        conn, pallet_id, "Закрыт после разбраковки",
        pallet["status"], "CLOSED",
        pallet["location"], pallet["location"],
        user, "Созданы: " + ", ".join(created_ids)
    )

    conn.commit()
    conn.close()

    return redirect(url_for("pallet_detail", pallet_id=pallet_id))


@app.route("/pallet/<pallet_id>/sold", methods=["POST"])
@shop5_required
def mark_sold(pallet_id: str):
    user = request.form.get("user", "").strip() or (current_user()["display_name"] if current_user() else "")
    comment = request.form.get("comment", "").strip()
    pallet = get_pallet_or_404(pallet_id)

    conn = get_db()
    conn.execute("""
        UPDATE pallets
        SET status = 'SOLD',
            location = 'SOLD',
            updated_at = ?
        WHERE id = ?
    """, (now_str(), pallet_id))

    add_movement(
        conn, pallet_id, "Продано",
        pallet["status"], "SOLD",
        pallet["location"], "SOLD",
        user, comment
    )

    conn.commit()
    conn.close()
    return redirect(url_for("pallet_detail", pallet_id=pallet_id))


@app.route("/pallet/<pallet_id>/label")
@login_required
def label(pallet_id: str):
    pallet = get_pallet_or_404(pallet_id)
    return render_template("label.html", pallet=pallet)


@app.route("/pallet/<pallet_id>/qr.svg")
@login_required
def qr_svg(pallet_id: str):
    # В QR кладём ссылку на карточку поддона.
    # SVG не требует Pillow, поэтому нормально ставится на Python 3.14.
    host_url = request.host_url.rstrip("/")
    data = f"{host_url}{url_for('pallet_detail', pallet_id=pallet_id)}"

    img = qrcode.make(data, image_factory=qrcode.image.svg.SvgImage)
    buf = io.BytesIO()
    img.save(buf)
    buf.seek(0)
    return Response(buf.getvalue(), mimetype="image/svg+xml")


@app.route("/export/stock.xlsx")
@shop5_required
def export_stock():
    conn = get_db()
    rows = conn.execute("""
        SELECT status, location, assortment,
               COUNT(*) AS pallets_count,
               SUM(rolls_count) AS rolls_sum,
               SUM(meters_total) AS meters_sum
        FROM pallets
        WHERE status NOT IN ('CLOSED', 'SOLD')
        GROUP BY status, location, assortment
        ORDER BY status, location, assortment
    """).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Остатки"

    headers = ["Статус", "Место", "Ассортимент", "Поддонов", "Рулонов", "Метраж"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=Side(style="thin"))

    for row in rows:
        ws.append([
            STATUSES.get(row["status"], row["status"]),
            LOCATIONS.get(row["location"], row["location"]),
            row["assortment"],
            row["pallets_count"],
            row["rolls_sum"] or 0,
            row["meters_sum"] or 0,
        ])

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"ostatki_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/assortments", methods=["GET", "POST"])
@shop5_required
def assortments_page():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        category = request.form.get("category", "fabric").strip()

        if name:
            conn = get_db()
            conn.execute("""
                INSERT OR REPLACE INTO assortments(name, category, active, created_at)
                VALUES (?, ?, 1, COALESCE((SELECT created_at FROM assortments WHERE name = ?), ?))
            """, (name, category, name, now_str()))
            conn.commit()
            conn.close()

        return redirect(url_for("assortments_page"))

    conn = get_db()
    rows = conn.execute("""
        SELECT * FROM assortments
        ORDER BY active DESC, category, name
    """).fetchall()
    conn.close()

    return render_template("assortments.html", assortments=rows)


@app.route("/assortments/<path:name>/toggle", methods=["POST"])
@shop5_required
def toggle_assortment(name: str):
    conn = get_db()
    row = conn.execute("SELECT active FROM assortments WHERE name = ?", (name,)).fetchone()
    if row:
        new_active = 0 if row["active"] else 1
        conn.execute("UPDATE assortments SET active = ? WHERE name = ?", (new_active, name))
        conn.commit()
    conn.close()
    return redirect(url_for("assortments_page"))


@app.route("/reports")
@shop5_required
def reports_page():
    month = request.args.get("month", datetime.now().strftime("%Y-%m"))

    conn = get_db()

    received = conn.execute("""
        SELECT p.source_shop,
               p.assortment,
               COUNT(*) AS pallets_count,
               SUM(p.rolls_count) AS rolls_sum,
               SUM(p.meters_total) AS meters_sum
        FROM movements m
        JOIN pallets p ON p.id = m.pallet_id
        WHERE m.action = 'Приёмка в цех 5'
          AND substr(m.timestamp, 1, 7) = ?
        GROUP BY p.source_shop, p.assortment
        ORDER BY p.source_shop, p.assortment
    """, (month,)).fetchall()

    processed = conn.execute("""
        SELECT assortment,
               process_type,
               machine,
               COUNT(*) AS pallets_count,
               SUM(rolls_count) AS rolls_sum,
               SUM(meters_total) AS meters_sum
        FROM pallets
        WHERE parent_id IS NOT NULL
          AND processing_date IS NOT NULL
          AND substr(created_at, 1, 7) = ?
        GROUP BY assortment, process_type, machine
        ORDER BY assortment, process_type, machine
    """, (month,)).fetchall()

    inspection_out = conn.execute("""
        SELECT status,
               location,
               assortment,
               COUNT(*) AS pallets_count,
               SUM(rolls_count) AS rolls_sum,
               SUM(meters_total) AS meters_sum
        FROM pallets
        WHERE parent_id IS NOT NULL
          AND status IN ('TO', 'WAREHOUSE', 'FG_WAREHOUSE', 'SHOP8_WAREHOUSE', 'LAB7', 'REJECT', 'CONDITIONALLY_OK')
          AND substr(created_at, 1, 7) = ?
        GROUP BY status, location, assortment
        ORDER BY status, location, assortment
    """, (month,)).fetchall()

    stock_by_warehouse = conn.execute("""
        SELECT location,
               assortment,
               COUNT(*) AS pallets_count,
               SUM(rolls_count) AS rolls_sum,
               SUM(meters_total) AS meters_sum
        FROM pallets
        WHERE status NOT IN ('CLOSED', 'SOLD')
          AND location IN ('SHOP5_WAREHOUSE', 'WAREHOUSE', 'FG_WAREHOUSE', 'SHOP8_WAREHOUSE')
        GROUP BY location, assortment
        ORDER BY location, assortment
    """).fetchall()

    stock = conn.execute("""
        SELECT status,
               location,
               assortment,
               COUNT(*) AS pallets_count,
               SUM(rolls_count) AS rolls_sum,
               SUM(meters_total) AS meters_sum
        FROM pallets
        WHERE status NOT IN ('CLOSED', 'SOLD')
        GROUP BY status, location, assortment
        ORDER BY location, status, assortment
    """).fetchall()

    sold = conn.execute("""
        SELECT assortment,
               COUNT(*) AS pallets_count,
               SUM(rolls_count) AS rolls_sum,
               SUM(meters_total) AS meters_sum
        FROM pallets
        WHERE status = 'SOLD'
          AND substr(updated_at, 1, 7) = ?
        GROUP BY assortment
        ORDER BY assortment
    """, (month,)).fetchall()

    conn.close()

    return render_template(
        "reports.html",
        month=month,
        received=received,
        processed=processed,
        inspection_out=inspection_out,
        stock_by_warehouse=stock_by_warehouse,
        stock=stock,
        sold=sold,
    )


@app.route("/export/monthly.xlsx")
@shop5_required
def export_monthly_report():
    month = request.args.get("month", datetime.now().strftime("%Y-%m"))

    conn = get_db()

    received = conn.execute("""
        SELECT p.source_shop, p.assortment,
               COUNT(*) AS pallets_count,
               SUM(p.rolls_count) AS rolls_sum,
               SUM(p.meters_total) AS meters_sum
        FROM movements m
        JOIN pallets p ON p.id = m.pallet_id
        WHERE m.action = 'Приёмка в цех 5'
          AND substr(m.timestamp, 1, 7) = ?
        GROUP BY p.source_shop, p.assortment
        ORDER BY p.source_shop, p.assortment
    """, (month,)).fetchall()

    processed = conn.execute("""
        SELECT assortment, process_type, machine,
               COUNT(*) AS pallets_count,
               SUM(rolls_count) AS rolls_sum,
               SUM(meters_total) AS meters_sum
        FROM pallets
        WHERE parent_id IS NOT NULL
          AND processing_date IS NOT NULL
          AND substr(created_at, 1, 7) = ?
        GROUP BY assortment, process_type, machine
        ORDER BY assortment, process_type, machine
    """, (month,)).fetchall()

    inspection_out = conn.execute("""
        SELECT status, location, assortment,
               COUNT(*) AS pallets_count,
               SUM(rolls_count) AS rolls_sum,
               SUM(meters_total) AS meters_sum
        FROM pallets
        WHERE parent_id IS NOT NULL
          AND status IN ('TO', 'WAREHOUSE', 'FG_WAREHOUSE', 'SHOP8_WAREHOUSE', 'LAB7', 'REJECT', 'CONDITIONALLY_OK')
          AND substr(created_at, 1, 7) = ?
        GROUP BY status, location, assortment
        ORDER BY status, location, assortment
    """, (month,)).fetchall()

    stock = conn.execute("""
        SELECT status, location, assortment,
               COUNT(*) AS pallets_count,
               SUM(rolls_count) AS rolls_sum,
               SUM(meters_total) AS meters_sum
        FROM pallets
        WHERE status NOT IN ('CLOSED', 'SOLD')
        GROUP BY status, location, assortment
        ORDER BY location, status, assortment
    """).fetchall()

    conn.close()

    wb = Workbook()

    def style_header(ws):
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(bottom=Side(style="thin"))
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 24

    ws = wb.active
    ws.title = "Поступило в цех 5"
    ws.append(["Цех-источник", "Ассортимент", "Поддонов", "Рулонов", "Метраж"])
    for r in received:
        ws.append([
            LOCATIONS.get(r["source_shop"], r["source_shop"]),
            r["assortment"],
            r["pallets_count"],
            r["rolls_sum"] or 0,
            r["meters_sum"] or 0,
        ])
    style_header(ws)

    ws = wb.create_sheet("Обработано")
    ws.append(["Ассортимент", "Вид обработки", "Установка", "Поддонов", "Рулонов", "Метраж"])
    for r in processed:
        ws.append([
            r["assortment"],
            r["process_type"] or "",
            LOCATIONS.get(r["machine"], r["machine"] or ""),
            r["pallets_count"],
            r["rolls_sum"] or 0,
            r["meters_sum"] or 0,
        ])
    style_header(ws)

    ws = wb.create_sheet("После разбраковки")
    ws.append(["Статус", "Место", "Ассортимент", "Поддонов", "Рулонов", "Метраж"])
    for r in inspection_out:
        ws.append([
            STATUSES.get(r["status"], r["status"]),
            LOCATIONS.get(r["location"], r["location"]),
            r["assortment"],
            r["pallets_count"],
            r["rolls_sum"] or 0,
            r["meters_sum"] or 0,
        ])
    style_header(ws)

    ws = wb.create_sheet("Остатки")
    ws.append(["Статус", "Место", "Ассортимент", "Поддонов", "Рулонов", "Метраж"])
    for r in stock:
        ws.append([
            STATUSES.get(r["status"], r["status"]),
            LOCATIONS.get(r["location"], r["location"]),
            r["assortment"],
            r["pallets_count"],
            r["rolls_sum"] or 0,
            r["meters_sum"] or 0,
        ])
    style_header(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"monthly_report_{month}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/search")
@login_required
def search():
    q = request.args.get("q", "").strip()
    user = current_user()
    conn = get_db()
    pallets = []
    if q:
        if is_source_user(user):
            pallets = conn.execute("""
                SELECT * FROM pallets
                WHERE source_shop = ?
                  AND parent_id IS NULL
                  AND COALESCE(created_reason, 'production_transfer') = 'production_transfer'
                  AND (
                       id LIKE ?
                    OR assortment LIKE ?
                    OR party_number LIKE ?
                    OR responsible LIKE ?
                    OR comment LIKE ?
                  )
                ORDER BY created_at DESC
                LIMIT 100
            """, (user["shop"], f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%")).fetchall()
        else:
            pallets = conn.execute("""
                SELECT * FROM pallets
                WHERE id LIKE ?
                   OR assortment LIKE ?
                   OR party_number LIKE ?
                   OR responsible LIKE ?
                   OR comment LIKE ?
                ORDER BY updated_at DESC
                LIMIT 100
            """, (f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%")).fetchall()
    conn.close()
    return render_template("search.html", q=q, pallets=pallets)


@app.route("/scan")
@login_required
def scan_page():
    return render_template("scan.html")


@app.route("/exp/shop10/roll-labels", methods=["GET", "POST"])
@login_required
def experimental_roll_labels():
    user = current_user()
    if not is_shop10_user(user):
        return render_template("access_denied.html", message="Экспериментальная вкладка доступна только цеху 10."), 403

    conn = get_db()

    if request.method == "POST":
        action = request.form.get("action", "create_roll")
        raw_loom_number = (request.form.get("loom_number") or "").strip()
        loom_number = int(raw_loom_number) if raw_loom_number.isdigit() else None
        if loom_number is not None and not (1 <= loom_number <= 8):
            conn.close()
            return render_template("access_denied.html", message="Номер станка должен быть от 1 до 8."), 400
        if action == "create_roll":
            roll_id = next_pallet_id("RL")
            conn.execute("""
                INSERT INTO roll_labels(
                    id, source_shop, assortment, roll_number, party_number, base_number, lubricant,
                    width, meters, meters_shift_1, meters_shift_2, weaver_name, weaver_name_2, loom_number,
                    assistant_name, production_date, created_by, created_at, comment
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                roll_id,
                user["shop"],
                request.form.get("assortment", "").strip(),
                request.form.get("roll_number", "").strip(),
                request.form.get("party_number", "").strip(),
                request.form.get("base_number", "").strip(),
                request.form.get("lubricant", "13").strip() or "13",
                request.form.get("width", "").strip(),
                parse_float(request.form.get("meters"), 0),
                parse_float(request.form.get("meters_shift_1"), 0),
                parse_float(request.form.get("meters_shift_2"), 0),
                request.form.get("weaver_name", "").strip(),
                request.form.get("weaver_name_2", "").strip(),
                loom_number,
                request.form.get("assistant_name", "").strip(),
                request.form.get("production_date") or "",
                user["display_name"],
                now_str(),
                request.form.get("comment", "").strip(),
            ))
        elif action == "build_pallet":
            selected_roll_ids = request.form.getlist("roll_ids")
            if selected_roll_ids:
                conn.execute("BEGIN IMMEDIATE")
                placeholders = ",".join(["?"] * len(selected_roll_ids))
                rows = conn.execute(
                    f"SELECT * FROM roll_labels WHERE id IN ({placeholders}) AND source_shop = ? AND pallet_id IS NULL",
                    (*selected_roll_ids, user["shop"])
                ).fetchall()
                if rows:
                    signature_pairs = {(r["assortment"], (r["party_number"] or "")) for r in rows}
                    if len(signature_pairs) != 1:
                        conn.rollback()
                        conn.close()
                        return render_template(
                            "access_denied.html",
                            message="Нельзя собрать один поддон из рулонов с разными ассортиментом или номером партии."
                        ), 400

                    rows = sorted(rows, key=lambda r: r["id"])
                    approved_roll_ids = [r["id"] for r in rows]
                    approved_placeholders = ",".join(["?"] * len(approved_roll_ids))
                    pallet_id = next_pallet_id_tx(conn, "P")
                    total_meters = sum(parse_float(r["meters"], 0) for r in rows)
                    assortment = rows[0]["assortment"]
                    party_number = rows[0]["party_number"] or ""
                    responsible = user["display_name"]
                    conn.execute("""
                        INSERT INTO pallets(
                            id, parent_id, source_shop, assortment, party_number, item_category, created_reason, rolls_count, meters_total,
                            production_date, processing_date, process_type, machine, responsible, status, location, created_by,
                            created_at, updated_at, comment
                        )
                        VALUES (?, NULL, ?, ?, ?, 'fabric', 'production_transfer', ?, ?, ?, NULL, NULL, NULL, ?, 'CREATED', ?, ?, ?, ?, ?)
                    """, (
                        pallet_id, user["shop"], assortment, party_number, len(rows), total_meters,
                        rows[0]["production_date"] or "", responsible, user["shop"], responsible,
                        now_str(), now_str(), f"Собран из рулонов: {', '.join(approved_roll_ids)}"
                    ))
                    add_movement(conn, pallet_id, "Сборка поддона из рулонов", None, "CREATED", None, user["shop"], responsible)
                    conn.execute(
                        f"UPDATE roll_labels SET pallet_id = ? WHERE id IN ({approved_placeholders})",
                        (pallet_id, *approved_roll_ids)
                    )

        elif action in ("update_roll", "delete_roll"):
            roll_id = request.form.get("roll_id", "").strip()
            roll = conn.execute("SELECT * FROM roll_labels WHERE id = ?", (roll_id,)).fetchone()
            if not can_manage_roll_label(user, roll):
                conn.rollback()
                conn.close()
                return render_template("access_denied.html", message="Редактирование/удаление доступно только в течение 1 дня для свободных рулонов цеха 10."), 403
            if action == "delete_roll":
                conn.execute("DELETE FROM roll_labels WHERE id = ?", (roll_id,))
            else:
                conn.execute("""
                    UPDATE roll_labels
                    SET assortment = ?, roll_number = ?, party_number = ?, base_number = ?, lubricant = ?,
                        width = ?, meters = ?, meters_shift_1 = ?, meters_shift_2 = ?, weaver_name = ?, weaver_name_2 = ?,
                        loom_number = ?, assistant_name = ?, production_date = ?, comment = ?
                    WHERE id = ?
                """, (
                    request.form.get("assortment", "").strip(),
                    request.form.get("roll_number", "").strip(),
                    request.form.get("party_number", "").strip(),
                    request.form.get("base_number", "").strip(),
                    request.form.get("lubricant", "13").strip() or "13",
                    request.form.get("width", "").strip(),
                    parse_float(request.form.get("meters"), 0),
                    parse_float(request.form.get("meters_shift_1"), 0),
                    parse_float(request.form.get("meters_shift_2"), 0),
                    request.form.get("weaver_name", "").strip(),
                    request.form.get("weaver_name_2", "").strip(),
                    loom_number,
                    request.form.get("assistant_name", "").strip(),
                    request.form.get("production_date") or "",
                    request.form.get("comment", "").strip(),
                    roll_id,
                ))
        conn.commit()
        conn.close()
        return redirect(url_for("experimental_roll_labels"))

    edit_roll_id = request.args.get("edit", "").strip()
    edit_roll = None
    if edit_roll_id:
        edit_roll = conn.execute("SELECT * FROM roll_labels WHERE id = ? AND source_shop = ?", (edit_roll_id, user["shop"])).fetchone()

    roll_labels = conn.execute("""
        SELECT * FROM roll_labels
        WHERE source_shop = ?
        ORDER BY created_at DESC
        LIMIT 200
    """, (user["shop"],)).fetchall()
    weavers = conn.execute(
        "SELECT full_name FROM shop10_staff WHERE role = 'weaver' AND active = 1 ORDER BY full_name"
    ).fetchall()
    assistants = conn.execute(
        "SELECT full_name FROM shop10_staff WHERE role = 'assistant' AND active = 1 ORDER BY full_name"
    ).fetchall()
    conn.close()
    return render_template(
        "roll_labels_experimental.html",
        roll_labels=roll_labels,
        edit_roll=edit_roll,
        weavers=weavers,
        assistants=assistants,
    )


@app.route("/exp/shop10/staff", methods=["GET", "POST"])
@login_required
def shop10_staff_page():
    user = current_user()
    if not is_shop10_user(user):
        return render_template("access_denied.html", message="Вкладка персонала доступна только цеху 10."), 403

    conn = get_db()
    if request.method == "POST":
        action = request.form.get("action", "").strip()
        role = request.form.get("role", "").strip()
        full_name = request.form.get("full_name", "").strip()
        if role in ("weaver", "assistant") and full_name:
            if action == "add_staff":
                conn.execute(
                    "INSERT OR IGNORE INTO shop10_staff(role, full_name, active, created_at) VALUES (?, ?, 1, ?)",
                    (role, full_name, now_str()),
                )
            elif action == "delete_staff":
                conn.execute("DELETE FROM shop10_staff WHERE role = ? AND full_name = ?", (role, full_name))
        conn.commit()
        conn.close()
        return redirect(url_for("shop10_staff_page"))

    staff = conn.execute(
        "SELECT role, full_name FROM shop10_staff WHERE active = 1 ORDER BY role, full_name"
    ).fetchall()
    conn.close()
    return render_template("shop10_staff.html", staff=staff)


@app.route("/roll-label/<roll_id>/qr.svg")
@login_required
def roll_qr_svg(roll_id: str):
    user = current_user()
    conn = get_db()
    roll = conn.execute("SELECT * FROM roll_labels WHERE id = ?", (roll_id,)).fetchone()
    conn.close()
    if not can_view_roll_label(user, roll):
        from flask import abort
        abort(404)

    host_url = request.host_url.rstrip("/")
    payload = f"{host_url}{url_for('roll_label_detail', roll_id=roll_id)}"

    img = qrcode.make(payload, image_factory=qrcode.image.svg.SvgImage)
    buf = io.BytesIO()
    img.save(buf)
    buf.seek(0)
    return Response(buf.getvalue(), mimetype="image/svg+xml")


@app.route("/roll-label/<roll_id>")
@login_required
def roll_label_detail(roll_id: str):
    user = current_user()
    conn = get_db()
    roll = conn.execute("SELECT * FROM roll_labels WHERE id = ?", (roll_id,)).fetchone()
    conn.close()
    if not can_view_roll_label(user, roll):
        from flask import abort
        abort(404)
    return render_template("roll_label_detail.html", roll=roll)


# Auto-init database on import for Gunicorn/systemd.
# This also applies lightweight SQLite migrations, for example party_number.
init_db()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8000, debug=True)
